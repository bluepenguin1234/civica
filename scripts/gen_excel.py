import csv, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

CSV = Path(__file__).parent.parent / "data" / "towns.csv"
OUT = Path(__file__).parent.parent / "civica_value_scores.xlsx"

MA_ZHVI = 613049.0
ZHVI = {
    "Cambridge":995293,"Lynn":537825,"Lawrence":455876,"Somerville":892143,
    "Haverhill":497568,"Medford":784462,"Peabody":652390,"Methuen":561211,
    "Arlington":1005584,"Salem":572734,"Woburn":704405,"Chelsea":505101,
    "Beverly":696077,"Andover":911128,"Lexington":1469802,"North Andover":760639,
    "Saugus":650624,"Danvers":674241,"Gloucester":693415,"Wakefield":756484,
    "Belmont":1377983,"Burlington":824833,"Reading":844034,"Winchester":1449988,
    "Newburyport":845604,"Amesbury":570013,"Marblehead":959425,"Uxbridge":487277,
    "Swampscott":763080,"Lynnfield":1018229,"Ipswich":790546,"Middleton":819582,
    "Salisbury":589785,"Georgetown":708796,"Boxford":989433,"Hamilton":827397,
    "Newbury":844611,"Groveland":648353,"Topsfield":898745,"Merrimac":597359,
    "Rockport":834071,"Rowley":733628,"Manchester-by-the-Sea":1183983,
    "Wenham":938834,"West Newbury":861236,"Essex":827029,"Nahant":903532,
    "Boston":720000,"Revere":440000,"Winthrop":520000,
    "Newton":1200000,"Waltham":720000,"Malden":545000,"Everett":530000,
    "Watertown":840000,"Framingham":575000,"Natick":775000,
    "Acton":840000,"Concord":1100000,"Stoneham":690000,
}
RATING_BANDS = [(82,"Great Value"),(65,"Good Value"),(50,"Fair Market"),(35,"Premium"),(0,"Luxury")]

def compute_value(score, name):
    zhvi = ZHVI.get(name)
    if not zhvi:
        return None, None
    raw = score / (zhvi / MA_ZHVI)
    vs = round(raw, 1)
    for t, label in RATING_BANDS:
        if vs >= t:
            return vs, label
    return vs, "Luxury"

# Load CSV towns
data = []
for r in csv.DictReader(open(CSV, encoding="utf-8")):
    if not r.get("civica_score"):
        continue
    vs, rating = compute_value(int(r["civica_score"]), r["town_name"])
    if vs is None:
        continue
    data.append({"town": r["town_name"], "score": int(r["civica_score"]),
                 "vs": vs, "rating": rating,
                 "zhvi": ZHVI.get(r["town_name"], ""), "county": r.get("county", "")})

data.sort(key=lambda x: x["vs"], reverse=True)
print(f"Total towns: {len(data)}")

# Build workbook
NAVY = "0D2D52"; WHITE = "FFFFFF"; LIGHT = "F0F6FF"; BORDER_COL = "DCE8FB"
rating_colors = {
    "Great Value": ("DCFCE7", "16A34A"),
    "Good Value":  ("DBEAFE", "1D4ED8"),
    "Fair Market": ("FEF9C3", "A16207"),
    "Premium":     ("FFE4E6", "BE123C"),
    "Luxury":      ("F3F4F6", "374151"),
}
thin = Side(style="thin", color=BORDER_COL)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Value Scores"

headers = ["Town", "Civica Score", "Value Score", "Value Rating", "ZHVI 2024", "County"]
ws.append(headers)
for col, h in enumerate(headers, 1):
    c = ws.cell(1, col)
    c.font = Font(bold=True, color=WHITE, size=11)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
ws.row_dimensions[1].height = 28

for i, d in enumerate(data):
    row = i + 2
    bg, fg = rating_colors.get(d["rating"], ("FFFFFF", "000000"))
    alt_bg = LIGHT if i % 2 == 0 else WHITE

    def wcell(col, val, fmt=None, align="center", row=row):
        c = ws.cell(row, col, val)
        c.border = border
        c.alignment = Alignment(horizontal=align, vertical="center")
        if fmt:
            c.number_format = fmt
        return c

    c = wcell(1, d["town"], align="left")
    c.font = Font(bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=alt_bg)

    c = wcell(2, d["score"])
    c.font = Font(bold=True, size=11, color=NAVY)
    c.fill = PatternFill("solid", fgColor=alt_bg)

    c = wcell(3, d["vs"], "#,##0.0")
    c.font = Font(bold=True, size=11, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)

    c = wcell(4, d["rating"])
    c.font = Font(bold=True, color=fg, size=9)
    c.fill = PatternFill("solid", fgColor=bg)

    zhvi_val = d["zhvi"] if d["zhvi"] else None
    c = wcell(5, zhvi_val, "$#,##0" if zhvi_val else None)
    c.fill = PatternFill("solid", fgColor=alt_bg)
    c.font = Font(size=9)

    c = wcell(6, d["county"], align="left")
    c.fill = PatternFill("solid", fgColor=alt_bg)
    c.font = Font(size=9)

    ws.row_dimensions[row].height = 20

ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 18
ws.freeze_panes = "A2"

# Legend sheet
ls = wb.create_sheet("Legend")
ls.column_dimensions["A"].width = 18
ls.column_dimensions["B"].width = 48
legend = [
    ("Rating",       "Meaning",                                              True),
    ("Great Value",  "Value Score >= 82  —  high quality, below-median price", False),
    ("Good Value",   "Value Score >= 65  —  solid quality for the price",       False),
    ("Fair Market",  "Value Score >= 50  —  quality priced in line with market",False),
    ("Premium",      "Value Score >= 35  —  paying up for quality",             False),
    ("Luxury",       "Value Score < 35   —  top-tier quality at top-tier price",False),
    ("",             "",                                                         False),
    ("Formula",      "Value Score = Civica Score / (Town ZHVI / MA State ZHVI)",False),
    ("MA ZHVI 2024", "$613,049  (Zillow Jan-2024 statewide median)",            False),
]
for li, (a, b, hdr) in enumerate(legend, 1):
    ca, cb = ls.cell(li, 1, a), ls.cell(li, 2, b)
    if hdr:
        ca.font = cb.font = Font(bold=True, color=WHITE)
        ca.fill = cb.fill = PatternFill("solid", fgColor=NAVY)
    elif a in rating_colors:
        bg, fg = rating_colors[a]
        ca.fill = cb.fill = PatternFill("solid", fgColor=bg)
        ca.font = cb.font = Font(bold=True, color=fg)
    ca.alignment = cb.alignment = Alignment(vertical="center")
    ls.row_dimensions[li].height = 20

wb.save(OUT)
print(f"Saved: {OUT}")
